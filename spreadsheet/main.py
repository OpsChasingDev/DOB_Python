import openpyxl

# read spreadsheet contents and store them
file = openpyxl.load_workbook("inventory.xlsx")
contents = file["Sheet1"]

# define empty dictionaries to populate later during iteration
product_per_supplier = {}
total_value_per_supplier = {}
products_under_10 = {}

# iterate through the rows in the spreadsheet, ignoring the header row
for row in range(2, (contents.max_row + 1)):
    supplier_name = contents.cell(row, 4).value
    inventory = contents.cell(row, 2).value
    price = contents.cell(row, 3).value
    product_num = contents.cell(row, 1).value
    inventory_price = contents.cell(row, 5)

    # calculate number of products per supplier
    if supplier_name in product_per_supplier:
        current_num_prod = product_per_supplier[supplier_name]
        product_per_supplier[supplier_name] = current_num_prod + 1
    else:
        print("adding new supplier")
        product_per_supplier[supplier_name] = 1

    # calculate total value of each supplier's inventory
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        print("adding new starting value of inventory")
        total_value_per_supplier[supplier_name] = inventory * price

    # save products whose inventory is less than 10
    if inventory < 10:
        products_under_10[int(product_num)] = int(inventory)

    # add value for total inventory price
    inventory_price.value = inventory * price

# show all logic values from dictionaries
print(product_per_supplier)
print(total_value_per_supplier)
print(products_under_10)

# create new file with additional info in column 5
file.save("inventory_with_total_value.xlsx")

